# -*- coding: utf-8 -*-
"""
GENERAR EXCEL de precios · myrenting.es
Vuelca TODAS las combinaciones modelo × km/año × meses × precio, por gestora,
a un .xlsx (una fila por combinación).

Fuentes (con fallback, para que funcione aquí y en el Mac):
  - M Automoción : data/raw/mautomocion.json  ->  mautomocion-db.json  ->  ofertas-manuales.json
                   (necesita la matriz km×meses: scrape con `run_weekly.py --con-matrix`)
  - Quadis + Kia : data/master/ofertas-manuales.json  (Quadis = PDF oficial, Kia = manual)

Salida: myrenting-precios.xlsx  (hojas Resumen · Precios · Fichas · una por gestora)

Uso:  python3 generar_excel.py
"""
import json, os, re, unicodedata
import pathlib as _pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = _pl.Path(__file__).resolve().parent

def load(p):
    p = REPO / p
    return json.load(open(p, encoding="utf-8")) if p.exists() else None

def first_existing(cands):
    for c in cands:
        d = load(c)
        if d is not None:
            return d, c
    return [], None

# --- 1. reunir ofertas de cada gestora ---
mauto_raw, mauto_src = first_existing(
    ["data/raw/mautomocion.json", "mautomocion-db.json", "ofertas-manuales.json"])
mauto = [o for o in mauto_raw if o.get("fuente") == "M Automoción"]

master = load("data/master/ofertas-manuales.json") or []
quadis_kia = [o for o in master if o.get("fuente") != "M Automoción"]

# detalle de Quadis (specs + FAQ) scrapeado aparte -> {MAKE|MODEL: {especificaciones, faq}}
quadis_det = load("data/master/quadis-detalle.json") or {}

def _nrm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper()
    s = s.replace(" BENZ", "").replace("MERCEDES-", "MERCEDES ")
    return re.sub(r"[^A-Z0-9]+", " ", s).strip()

# índice normalizado de Quadis: clave "MARCA|PRIMEROS 2 TOKENS DEL MODELO"
quadis_det_norm = {}
for k, v in quadis_det.items():
    mk, _, md = k.partition("|")
    nk = _nrm(mk) + "|" + " ".join(_nrm(md).split()[:2])
    quadis_det_norm.setdefault(nk, v)

def quadis_detalle_de(make, model):
    exact = quadis_det.get(f"{(make or '').upper()}|{(model or '').upper()}")
    if exact:
        return exact
    nk = _nrm(make) + "|" + " ".join(_nrm(model).split()[:2])
    return quadis_det_norm.get(nk)

ofertas = mauto + quadis_kia
print(f"M Automoción: {len(mauto)} ofertas (de {mauto_src})")
print(f"Quadis + Kia: {len(quadis_kia)} ofertas (de data/master/ofertas-manuales.json)")

# --- 2. aplanar a filas: una por combinación km×meses ---
TIPO_LBL = {"particular": "Particular", "autonomo": "Autónomo", "empresa": "Empresa"}
def iva_lbl(tipo):
    return "IVA incluido" if tipo == "particular" else "+ IVA (neto)"

filas = []
for o in ofertas:
    marca = (o.get("make") or "").title()
    modelo = (o.get("model") or "").title()
    tipo = o.get("tipo") or ""
    base = dict(
        gestora=o.get("fuente") or "",
        tipo=TIPO_LBL.get(tipo, tipo.title()),
        marca=marca,
        modelo=modelo,
        version=o.get("version") or "",
        combustible=o.get("fuel") or "",
        categoria=o.get("category") or "",
        iva=iva_lbl(tipo),
        url=o.get("url") or "",
    )
    combos = o.get("combinaciones") or []
    if combos:
        for c in combos:
            filas.append({**base, "km": c.get("km"), "meses": c.get("meses"), "precio": c.get("precio")})
    else:
        # sin matriz: una sola fila con el precio "desde"
        filas.append({**base, "km": o.get("km"), "meses": o.get("duracion"), "precio": o.get("precio_desde")})

# orden estable: gestora, marca, modelo, tipo, km, meses
filas.sort(key=lambda f: (f["gestora"], f["marca"], f["modelo"], f["tipo"],
                          f["km"] or 0, f["meses"] or 0))
print(f"Total filas (combinaciones): {len(filas)}")

# --- 2b. fichas: una fila por vehículo con TODO el detalle (si el scrape lo trae) ---
def esp(det, *claves):
    e = (det or {}).get("especificaciones", {}) or {}
    low = {str(k).strip().lower(): v for k, v in e.items()}
    for c in claves:
        v = low.get(c.strip().lower())
        if v:
            return v
    return ""

def faq_txt(det):
    fs = (det or {}).get("faq", []) or []
    return "  |  ".join(f"{x.get('q','')} {x.get('a','')}".strip() for x in fs)

def fila_ficha(o):
    # M Automoción trae 'detalle' en la oferta; Quadis lo trae del mapa por MAKE|MODEL
    det = o.get("detalle")
    if not det and o.get("fuente") == "Quadis":
        det = quadis_detalle_de(o.get("make"), o.get("model"))
    return {
        "gestora": o.get("fuente") or "",
        "tipo": TIPO_LBL.get(o.get("tipo") or "", (o.get("tipo") or "").title()),
        "marca": (o.get("make") or "").title(),
        "modelo": (o.get("model") or "").title(),
        "version": o.get("version") or "",
        # etiquetas de M Automoción (MAYÚSCULAS) y de Quadis (Capitalizadas) a la vez
        "combustible": esp(det, "COMBUSTIBLE", "Combustible") or o.get("fuel") or "",
        "transmision": esp(det, "TRANSMISIÓN", "TRANSMISION", "Cambio"),
        "etiqueta": esp(det, "ETIQUETA", "Etiqueta", "Distintivo"),
        "motor": esp(det, "MOTOR", "Potencia"),
        "estado": esp(det, "ESTADO", "Estado"),
        "carroceria": esp(det, "CARROCERÍA", "CARROCERIA", "Carrocería") or o.get("category") or "",
        "traccion": esp(det, "TRACCIÓN", "TRACCION", "Tracción"),
        "puertas": esp(det, "PUERTAS", "Puertas"),
        "plazas": esp(det, "PLAZAS", "Plazas"),
        "color": esp(det, "COLOR", "Color"),
        "acabado": esp(det, "ACABADO", "Acabado"),
        "consumo": esp(det, "CONSUMO", "Consumo"),
        "equip_ext": " · ".join((det or {}).get("equip_exterior", [])),
        "equip_int": " · ".join((det or {}).get("equip_interior", [])),
        "coberturas": (det or {}).get("coberturas", ""),
        "notas": (det or {}).get("notas", ""),
        "faq": faq_txt(det),
        "url": (o.get("url") or "").split("#")[0],
    }

# por (gestora,marca,modelo,tipo) elige la oferta CON detalle si la hay
best = {}
for o in ofertas:
    clave = (o.get("fuente"), o.get("make"), o.get("model"), o.get("tipo"))
    cur = best.get(clave)
    if cur is None or (o.get("detalle") and not cur.get("detalle")):
        best[clave] = o
fichas = [fila_ficha(o) for o in best.values()]
fichas.sort(key=lambda f: (f["gestora"], f["marca"], f["modelo"], f["tipo"]))
con_detalle = sum(1 for f in fichas if f["equip_int"] or f["motor"])
print(f"Total fichas (vehículos): {len(fichas)} ({con_detalle} con detalle completo)")

# --- 3. escribir el .xlsx ---
COLS = [
    ("Gestora", "gestora", 15),
    ("Tipo cliente", "tipo", 13),
    ("Marca", "marca", 16),
    ("Modelo", "modelo", 26),
    ("Versión", "version", 46),
    ("Combustible", "combustible", 18),
    ("Categoría", "categoria", 14),
    ("Km/año", "km", 10),
    ("Meses", "meses", 8),
    ("Precio €/mes", "precio", 13),
    ("IVA", "iva", 14),
    ("URL", "url", 40),
]
HEAD_FILL = PatternFill("solid", fgColor="F04E00")
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="E5E5E5")
BORDER = Border(bottom=THIN)

def escribir_hoja(ws, rows, cols=COLS):
    # cabecera
    for ci, (titulo, _, ancho) in enumerate(cols, 1):
        cell = ws.cell(1, ci, titulo)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = ancho
    # datos
    for ri, f in enumerate(rows, 2):
        for ci, (_, key, _) in enumerate(cols, 1):
            val = f.get(key)
            cell = ws.cell(ri, ci, val)
            cell.font = CELL_FONT
            cell.border = BORDER
            if key == "precio" and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00 "€"'
            elif key == "km" and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            if key in ("km", "meses", "precio", "puertas", "plazas"):
                cell.alignment = Alignment(horizontal="center")
            elif key in ("equip_ext", "equip_int", "coberturas", "notas", "faq"):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

FICHA_COLS = [
    ("Gestora", "gestora", 15), ("Tipo cliente", "tipo", 13),
    ("Marca", "marca", 16), ("Modelo", "modelo", 24), ("Versión", "version", 40),
    ("Combustible", "combustible", 18), ("Transmisión", "transmision", 14),
    ("Etiqueta", "etiqueta", 10), ("Motor", "motor", 10), ("Estado", "estado", 10),
    ("Carrocería", "carroceria", 13), ("Tracción", "traccion", 16),
    ("Puertas", "puertas", 8), ("Plazas", "plazas", 8),
    ("Color", "color", 18), ("Acabado", "acabado", 14), ("Consumo", "consumo", 14),
    ("Equip. exterior", "equip_ext", 55), ("Equip. interior", "equip_int", 65),
    ("Coberturas", "coberturas", 55), ("Notas", "notas", 55),
    ("FAQ", "faq", 60), ("URL", "url", 40),
]

wb = Workbook()
ws = wb.active; ws.title = "Precios"
escribir_hoja(ws, filas)
escribir_hoja(wb.create_sheet("Fichas"), fichas, FICHA_COLS)

# una hoja por gestora
gestoras = sorted({f["gestora"] for f in filas})
for g in gestoras:
    sub = [f for f in filas if f["gestora"] == g]
    nombre = g[:28].replace("/", "-") or "s-gestora"
    escribir_hoja(wb.create_sheet(nombre), sub)

# hoja Resumen (recuento por gestora)
rs = wb.create_sheet("Resumen")
rs["A1"] = "Gestora"; rs["B1"] = "Combinaciones"
for c in ("A1", "B1"):
    rs[c].fill = HEAD_FILL; rs[c].font = HEAD_FONT
    rs[c].alignment = Alignment(horizontal="center")
from collections import Counter
por_gestora = Counter(f["gestora"] for f in filas)
for i, g in enumerate(gestoras, 2):
    rs.cell(i, 1, g).font = CELL_FONT
    rs.cell(i, 2, por_gestora[g]).font = CELL_FONT
tot = len(gestoras) + 2
rs.cell(tot, 1, "TOTAL").font = Font(name="Arial", bold=True)
rs.cell(tot, 2, len(filas)).font = Font(name="Arial", bold=True)
rs.column_dimensions["A"].width = 18; rs.column_dimensions["B"].width = 16
wb.move_sheet("Resumen", -(len(wb.sheetnames)-1))  # Resumen al principio

OUT = REPO / "myrenting-precios.xlsx"
wb.save(OUT)
print(f"✅ {OUT.name}: {len(filas)} filas, {len(gestoras)} gestoras, {len(wb.sheetnames)} hojas")
